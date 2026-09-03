from __future__ import annotations

from datetime import date, datetime, time
import json
import math
from pathlib import Path
from typing import Any, TextIO
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from .readable_converter import ReadableConverter


class ExcelConverter(ReadableConverter):
    """
    Convert Office Open XML Excel workbooks to bounded UTF-8 text for LLMs.

    Supported formats
    -----------------
    ``.xlsx``, ``.xlsm``, ``.xltx``, and ``.xltm`` are ZIP-based Office Open
    XML / SpreadsheetML packages. Legacy BIFF ``.xls`` and binary ``.xlsb``
    are different formats and intentionally are not handled here.

    Format model
    ------------
    An Office Open XML document is an OPC package: a ZIP archive containing
    typed parts connected by relationships. A SpreadsheetML workbook part
    references worksheet parts; worksheet ``sheetData`` contains rows/cells.
    Cells can represent numbers, strings, booleans, errors, dates, or formulas.
    Repeated strings may be stored in a workbook-level shared-string table;
    openpyxl resolves that storage detail and returns the logical cell value.

    Formula handling is deliberately non-evaluating. SpreadsheetML stores the
    formula expression separately from an optional cached value produced by the
    spreadsheet application during its last calculation. openpyxl likewise
    exposes formulas with ``data_only=False`` and cached results with
    ``data_only=True``; it does not calculate formulas. This converter opens
    the workbook in both modes and reports cached results as cached, never as
    freshly calculated values. Cached results may be absent or stale.

    The text representation preserves LLM-relevant structure where openpyxl
    exposes it: sheet order/visibility, document and calculation metadata,
    defined names, merged ranges, tables, filters, freeze panes, print areas,
    hidden dimensions, validations, conditional-format ranges, comments,
    hyperlinks, number formats/styles, array formulas, and data-table formulas.
    Charts, images, VBA, embedded/OLE objects, and other visual/binary parts are
    not rendered or executed; their presence is only summarized.

    Safety / boundedness
    --------------------
    XLSX-family files are ZIP packages, so package entry count and declared
    uncompressed size are checked before parsing. Output is capped globally and
    per worksheet. The sparse-cell enumerator uses ``Worksheet._cells`` when
    available because openpyxl does not currently expose a public sparse-cell
    iterator; this avoids walking pathological dimensions such as
    A1:XFD1048576. That private dependency is isolated in ``_material_cells``
    and has a public-API fallback.

    References used for the format decisions above
    ------------------------------------------------
    [ECMA-376]
        ECMA-376, Office Open XML File Formats, especially Part 2 (Open
        Packaging Conventions) and SpreadsheetML material in Part 4.

    [ISO-29500]
        ISO/IEC 29500, Office Open XML File Formats.

    [MS-SPREADSHEET-STRUCTURE]
        Microsoft Learn, "Structure of a SpreadsheetML document".
        Documents workbook/worksheet parts, relationships, and sheetData.

    [MS-SHEETS]
        Microsoft Learn, "Working with sheets".
        Documents rows, cells, values, data types, and shared strings.

    [MS-FORMULAS]
        Microsoft Learn, "Working with formulas".
        Documents formula text and last-calculated cached cell values.

    [MS-SHARED-STRINGS]
        Microsoft Learn, "Working with the shared string table".
        Documents workbook-level shared-string storage and cell indexes.

    [OPENPYXL-LOAD]
        openpyxl 3.1, ``load_workbook`` documentation.
        ``data_only`` selects formula text vs cached values; ``read_only``
        reduces memory but omits features; ``rich_text`` preserves rich text.

    [OPENPYXL-FORMULAS]
        openpyxl 3.1 formula documentation.
        openpyxl preserves simple, array, and data-table formulas but does not
        evaluate them.

    [OPENPYXL-STRUCTURE]
        openpyxl 3.1 documentation for tables, defined names, merged ranges,
        filters, validations, worksheet metadata, and cell properties.
    """

    VERSION = 1

    MAX_SOURCE_BYTES = 128 * 1024 * 1024
    MAX_PACKAGE_ENTRIES = 20_000
    MAX_UNCOMPRESSED_BYTES = 768 * 1024 * 1024

    MAX_WORKSHEETS = 64
    MAX_CHARTSHEETS = 64
    MAX_TOTAL_CELLS = 50_000
    MAX_CELLS_PER_SHEET = 12_000
    MAX_CELL_CHARS = 4_000
    MAX_ANNOTATION_CHARS = 2_000

    MAX_DEFINED_NAMES = 256
    MAX_MERGED_RANGES = 256
    MAX_TABLES = 128
    MAX_TABLE_COLUMNS = 256
    MAX_DATA_VALIDATIONS = 128
    MAX_CONDITIONAL_FORMATS = 128
    MAX_HIDDEN_DIMENSIONS = 256
    MAX_SPECIAL_FORMULAS = 128

    FORMAT_REFERENCES = (
        "ECMA-376: Office Open XML File Formats",
        "ISO/IEC 29500: Office Open XML File Formats",
        "Microsoft Learn: Structure of a SpreadsheetML document",
        "Microsoft Learn: Working with sheets",
        "Microsoft Learn: Working with formulas",
        "Microsoft Learn: Working with the shared string table",
        "openpyxl 3.1: load_workbook",
        "openpyxl 3.1: formulas, tables, names, merges, validations",
    )

    @property
    def extensions(self) -> frozenset[str]:
        """Handle extensions."""
        return frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})

    def _convert(self, *, source: Path, destination: Path) -> None:
        """Handle convert."""
        package = self._inspect_package(source)

        # Normal mode is intentional: read-only mode omits or restricts some
        # worksheet features that this representation wants to preserve.
        formula_book = load_workbook(
            source,
            read_only=False,
            data_only=False,
            keep_vba=False,
            keep_links=True,
            rich_text=True,
        )
        cached_book = None

        try:
            cached_book = load_workbook(
                source,
                read_only=False,
                data_only=True,
                keep_vba=False,
                keep_links=False,
                rich_text=True,
            )

            with destination.open("w", encoding="utf-8", newline="\n") as output:
                self._write_header(output, source, formula_book, package)
                self._write_document_properties(output, formula_book)
                self._write_calculation_properties(output, formula_book)
                self._write_defined_names(output, formula_book)
                self._write_chartsheets(output, formula_book)

                worksheets = list(formula_book.worksheets)
                selected = worksheets[: self.MAX_WORKSHEETS]
                remaining_cells = self.MAX_TOTAL_CELLS

                for index, worksheet in enumerate(selected, 1):
                    cached_sheet = (
                        cached_book[worksheet.title]
                        if worksheet.title in cached_book.sheetnames
                        else None
                    )
                    used = self._write_sheet(
                        output,
                        index=index,
                        worksheet=worksheet,
                        cached_sheet=cached_sheet,
                        remaining_cells=remaining_cells,
                    )
                    remaining_cells -= used

                    if remaining_cells <= 0:
                        output.write(
                            "\n===== Workbook cell output truncated =====\n"
                            f"Global material-cell limit reached: "
                            f"{self.MAX_TOTAL_CELLS}.\n"
                        )
                        break

                if len(worksheets) > len(selected):
                    output.write(
                        "\n===== Worksheets omitted =====\n"
                        f"{len(worksheets) - len(selected)} worksheet(s) omitted "
                        f"after the {self.MAX_WORKSHEETS}-worksheet limit.\n"
                    )

                output.write(
                    "\n===== Interpretation notes =====\n"
                    "- Formula results labelled cached come from the workbook "
                    "and can be missing or stale.\n"
                    "- Formulas and VBA/macros are never executed.\n"
                    "- Number formats are reported, but Excel's full visual "
                    "rendering is not reproduced.\n"
                    "- Charts, images, OLE/embedded objects, slicers, and other "
                    "visual/binary artifacts are not rendered.\n"
                    "- Legacy .xls and binary .xlsb require other converters.\n"
                )
        finally:
            formula_book.close()
            if cached_book is not None:
                cached_book.close()

    def _inspect_package(self, source: Path) -> dict[str, Any]:
        """Handle inspect package."""
        source_size = source.stat().st_size
        if source_size > self.MAX_SOURCE_BYTES:
            raise ValueError(
                f"Excel file is too large: {source_size} bytes exceeds "
                f"{self.MAX_SOURCE_BYTES}."
            )

        try:
            with ZipFile(source) as archive:
                entries = archive.infolist()
                if len(entries) > self.MAX_PACKAGE_ENTRIES:
                    raise ValueError(
                        f"Excel package has {len(entries)} ZIP entries; limit is "
                        f"{self.MAX_PACKAGE_ENTRIES}."
                    )

                uncompressed = sum(entry.file_size for entry in entries)
                if uncompressed > self.MAX_UNCOMPRESSED_BYTES:
                    raise ValueError(
                        f"Excel package expands to {uncompressed} bytes; limit is "
                        f"{self.MAX_UNCOMPRESSED_BYTES}."
                    )

                names = {entry.filename for entry in entries}
                if "[Content_Types].xml" not in names:
                    raise ValueError(
                        "Invalid Office Open XML package: "
                        "[Content_Types].xml is missing."
                    )

                lower = {name.lower() for name in names}
                return {
                    "entries": len(entries),
                    "source_bytes": source_size,
                    "uncompressed_bytes": uncompressed,
                    "has_vba": any(n.endswith("vbaproject.bin") for n in lower),
                    "has_external_links": any(
                        n.startswith("xl/externallinks/") for n in lower
                    ),
                    "has_charts": any(n.startswith("xl/charts/") for n in lower),
                    "has_images": any(n.startswith("xl/media/") for n in lower),
                    "has_pivots": any(
                        n.startswith("xl/pivottables/")
                        or n.startswith("xl/pivotcache/")
                        for n in lower
                    ),
                    "has_embeddings": any(
                        n.startswith("xl/embeddings/") for n in lower
                    ),
                }
        except BadZipFile as error:
            raise ValueError(
                "Invalid Excel/Office Open XML file: not a valid ZIP package."
            ) from error

    def _write_header(
        self,
        output: TextIO,
        source: Path,
        workbook: Any,
        package: dict[str, Any],
    ) -> None:
        """Handle write header."""
        output.write(
            f"Excel Workbook: {source.name}\n"
            f"Format: {source.suffix.lower()} (Office Open XML / SpreadsheetML)\n"
            f"Worksheets: {len(workbook.worksheets)}\n"
            f"Chartsheets: {len(workbook.chartsheets)}\n"
            f"Sheet order: {', '.join(workbook.sheetnames)}\n"
            f"Package parts: {package['entries']}\n"
            f"Package size: {package['source_bytes']} bytes compressed; "
            f"{package['uncompressed_bytes']} bytes declared uncompressed\n"
            f"Workbook epoch: {self._render_scalar(workbook.epoch)}\n"
        )

        features: list[str] = []
        feature_flags = (
            ("has_vba", "VBA project (not executed)"),
            ("has_external_links", "external-link parts"),
            ("has_charts", "charts"),
            ("has_images", "images/media"),
            ("has_pivots", "pivot-table/cache parts"),
            ("has_embeddings", "embedded/OLE parts"),
        )
        for key, label in feature_flags:
            if package[key]:
                features.append(label)
        if features:
            output.write("Package features: " + ", ".join(features) + "\n")

        output.write(
            "Formula policy: formulas are preserved; cached results are read "
            "separately and never recalculated by this converter.\n"
        )

    def _write_document_properties(self, output: TextIO, workbook: Any) -> None:
        """Handle write document properties."""
        props = workbook.properties
        fields = (
            ("Title", props.title),
            ("Subject", props.subject),
            ("Creator", props.creator),
            ("Last modified by", props.lastModifiedBy),
            ("Created", props.created),
            ("Modified", props.modified),
            ("Category", props.category),
            ("Description", props.description),
            ("Keywords", props.keywords),
            ("Identifier", props.identifier),
            ("Language", props.language),
            ("Version", props.version),
            ("Revision", props.revision),
        )
        visible = [(name, value) for name, value in fields if value not in (None, "")]
        custom = getattr(workbook, "custom_doc_props", None)
        if not visible and not custom:
            return

        output.write("\n===== Document properties =====\n")
        for name, value in visible:
            output.write(f"{name}: {self._render_scalar(value)}\n")

        if custom:
            output.write("Custom properties:\n")
            for item in custom:
                name = getattr(item, "name", "(unnamed)")
                value = getattr(item, "value", None)
                output.write(f"- {name}: {self._render_scalar(value)}\n")

    def _write_calculation_properties(self, output: TextIO, workbook: Any) -> None:
        """Handle write calculation properties."""
        calc = getattr(workbook, "calculation", None)
        if calc is None:
            return

        fields = (
            ("mode", getattr(calc, "calcMode", None)),
            ("calculation id", getattr(calc, "calcId", None)),
            ("full calculation on load", getattr(calc, "fullCalcOnLoad", None)),
            ("force full calculation", getattr(calc, "forceFullCalc", None)),
            ("calculate on save", getattr(calc, "calcOnSave", None)),
            ("iterate", getattr(calc, "iterate", None)),
            ("iteration count", getattr(calc, "iterateCount", None)),
            ("iteration delta", getattr(calc, "iterateDelta", None)),
        )
        visible = [(name, value) for name, value in fields if value is not None]
        if not visible:
            return

        output.write("\n===== Calculation settings =====\n")
        for name, value in visible:
            output.write(f"{name}: {self._render_scalar(value)}\n")

    def _write_defined_names(self, output: TextIO, workbook: Any) -> None:
        """Handle write defined names."""
        names: list[tuple[str, Any]] = [
            ("workbook", item) for item in workbook.defined_names.values()
        ]
        for worksheet in workbook.worksheets:
            names.extend(
                (f"sheet:{worksheet.title}", item)
                for item in worksheet.defined_names.values()
            )

        if not names:
            return

        output.write("\n===== Defined names =====\n")
        for scope, item in names[: self.MAX_DEFINED_NAMES]:
            name = getattr(item, "name", "(unnamed)")
            value = getattr(item, "attr_text", None)
            if value is None:
                value = getattr(item, "value", "")

            annotations = [f"scope={scope}"]
            if getattr(item, "hidden", False):
                annotations.append("hidden")
            if getattr(item, "is_external", False):
                annotations.append("external")
            kind = getattr(item, "type", None)
            if kind:
                annotations.append(f"type={kind}")

            output.write(
                f"- {name}: {self._bounded_inline(value, self.MAX_ANNOTATION_CHARS)} "
                f"[{', '.join(annotations)}]\n"
            )

        if len(names) > self.MAX_DEFINED_NAMES:
            output.write(
                f"[Defined names truncated: "
                f"{len(names) - self.MAX_DEFINED_NAMES} omitted]\n"
            )

    def _write_chartsheets(self, output: TextIO, workbook: Any) -> None:
        """Handle write chartsheets."""
        chartsheets = list(workbook.chartsheets)
        if not chartsheets:
            return

        output.write("\n===== Chartsheets =====\n")
        for sheet in chartsheets[: self.MAX_CHARTSHEETS]:
            state = getattr(sheet, "sheet_state", "unknown")
            output.write(
                f"- {sheet.title} [state={state}; chart content not rendered]\n"
            )
        if len(chartsheets) > self.MAX_CHARTSHEETS:
            output.write(
                f"[Chartsheets truncated: "
                f"{len(chartsheets) - self.MAX_CHARTSHEETS} omitted]\n"
            )

    def _write_sheet(
        self,
        output: TextIO,
        *,
        index: int,
        worksheet: Any,
        cached_sheet: Any | None,
        remaining_cells: int,
    ) -> int:
        """Handle write sheet."""
        output.write(f"\n===== Worksheet {index}: {worksheet.title} =====\n")
        self._write_sheet_metadata(output, worksheet)
        self._write_sheet_structures(output, worksheet)

        cells = self._material_cells(worksheet)
        if not cells:
            output.write("\nCells: [no material cell values/comments/hyperlinks]\n")
            return 0

        min_row = min(cell.row for cell in cells)
        max_row = max(cell.row for cell in cells)
        min_col = min(cell.column for cell in cells)
        max_col = max(cell.column for cell in cells)
        output.write(
            "\nMaterial bounds: "
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}\n"
            f"Material cells: {len(cells)}\n"
            "Cells (row-wise sparse representation):\n"
        )

        allowed = min(len(cells), self.MAX_CELLS_PER_SHEET, remaining_cells)
        selected = cells[:allowed]
        current_row: int | None = None
        row_parts: list[str] = []

        for cell in selected:
            if current_row is None:
                current_row = cell.row
            if cell.row != current_row:
                output.write(f"{current_row}: " + " | ".join(row_parts) + "\n")
                current_row = cell.row
                row_parts = []

            cached_cell = cached_sheet[cell.coordinate] if cached_sheet is not None else None
            row_parts.append(
                f"{get_column_letter(cell.column)}: "
                + self._render_cell(cell, cached_cell)
            )

        if current_row is not None:
            output.write(f"{current_row}: " + " | ".join(row_parts) + "\n")

        if allowed < len(cells):
            output.write(
                f"[Cell output truncated: {len(cells) - allowed} material cell(s) omitted]\n"
            )
        return allowed

    def _write_sheet_metadata(self, output: TextIO, worksheet: Any) -> None:
        """Handle write sheet metadata."""
        output.write(
            f"State: {worksheet.sheet_state}\n"
            f"Stored dimension: {worksheet.calculate_dimension()}\n"
        )

        freeze = worksheet.freeze_panes
        if freeze:
            output.write(f"Freeze panes: {getattr(freeze, 'coordinate', freeze)}\n")

        if worksheet.auto_filter.ref:
            output.write(f"Auto filter: {worksheet.auto_filter.ref}\n")
        if worksheet.print_area:
            output.write(f"Print area: {worksheet.print_area}\n")
        if worksheet.print_title_rows:
            output.write(f"Print title rows: {worksheet.print_title_rows}\n")
        if worksheet.print_title_cols:
            output.write(f"Print title columns: {worksheet.print_title_cols}\n")
        if getattr(worksheet.protection, "sheet", False):
            output.write("Sheet protection: enabled\n")

        hidden_rows = [
            str(index)
            for index, dimension in worksheet.row_dimensions.items()
            if dimension.hidden
        ]
        hidden_cols = [
            str(key)
            for key, dimension in worksheet.column_dimensions.items()
            if dimension.hidden
        ]
        self._write_bounded_list(
            output, "Hidden rows", hidden_rows, self.MAX_HIDDEN_DIMENSIONS
        )
        self._write_bounded_list(
            output, "Hidden columns", hidden_cols, self.MAX_HIDDEN_DIMENSIONS
        )

    def _write_sheet_structures(self, output: TextIO, worksheet: Any) -> None:
        """Handle write sheet structures."""
        merged = sorted(
            (str(cell_range) for cell_range in worksheet.merged_cells.ranges),
            key=str.lower,
        )
        self._write_bounded_list(
            output, "Merged ranges", merged, self.MAX_MERGED_RANGES
        )
        self._write_tables(output, worksheet)
        self._write_data_validations(output, worksheet)
        self._write_conditional_formats(output, worksheet)

        array_formulae = getattr(worksheet, "array_formulae", {})
        if array_formulae:
            output.write("Array formulas:\n")
            items = list(array_formulae.items())
            for coordinate, ref in items[: self.MAX_SPECIAL_FORMULAS]:
                output.write(f"- {coordinate}: {ref}\n")
            if len(items) > self.MAX_SPECIAL_FORMULAS:
                output.write(
                    f"[Array formulas truncated: "
                    f"{len(items) - self.MAX_SPECIAL_FORMULAS} omitted]\n"
                )

        table_formulae = getattr(worksheet, "table_formulae", {})
        if table_formulae:
            output.write("Data-table formulas:\n")
            items = list(table_formulae.items())
            for coordinate, ref in items[: self.MAX_SPECIAL_FORMULAS]:
                output.write(f"- {coordinate}: {ref}\n")
            if len(items) > self.MAX_SPECIAL_FORMULAS:
                output.write(
                    f"[Data-table formulas truncated: "
                    f"{len(items) - self.MAX_SPECIAL_FORMULAS} omitted]\n"
                )

    def _write_tables(self, output: TextIO, worksheet: Any) -> None:
        """Handle write tables."""
        tables = list(worksheet.tables.values())
        if not tables:
            return

        output.write("Tables:\n")
        for table in tables[: self.MAX_TABLES]:
            name = table.displayName or table.name or "(unnamed)"
            type_suffix = f" [type={table.tableType}]" if table.tableType else ""
            output.write(f"- {name}: {table.ref}{type_suffix}\n")

            columns = list(table.tableColumns)
            if columns:
                rendered: list[str] = []
                for column in columns[: self.MAX_TABLE_COLUMNS]:
                    text = str(column.name)
                    calculated = column.calculatedColumnFormula
                    if calculated is not None and calculated.text:
                        text += (
                            " {formula="
                            + self._bounded_inline(
                                calculated.text, self.MAX_ANNOTATION_CHARS
                            )
                            + "}"
                        )
                    if column.totalsRowFunction:
                        text += f" {{total={column.totalsRowFunction}}}"
                    rendered.append(text)
                output.write("  columns: " + ", ".join(rendered) + "\n")
                if len(columns) > self.MAX_TABLE_COLUMNS:
                    output.write(
                        f"  [Table columns truncated: "
                        f"{len(columns) - self.MAX_TABLE_COLUMNS} omitted]\n"
                    )

        if len(tables) > self.MAX_TABLES:
            output.write(
                f"[Tables truncated: {len(tables) - self.MAX_TABLES} omitted]\n"
            )

    def _write_data_validations(self, output: TextIO, worksheet: Any) -> None:
        """Handle write data validations."""
        collection = getattr(worksheet, "data_validations", None)
        validations = list(getattr(collection, "dataValidation", ()))
        if not validations:
            return

        output.write("Data validations:\n")
        for validation in validations[: self.MAX_DATA_VALIDATIONS]:
            parts = [
                f"range={validation.sqref}",
                f"type={validation.type or 'custom/unspecified'}",
            ]
            if validation.operator:
                parts.append(f"operator={validation.operator}")
            if validation.formula1 is not None:
                parts.append(
                    "formula1="
                    + self._bounded_inline(
                        validation.formula1, self.MAX_ANNOTATION_CHARS
                    )
                )
            if validation.formula2 is not None:
                parts.append(
                    "formula2="
                    + self._bounded_inline(
                        validation.formula2, self.MAX_ANNOTATION_CHARS
                    )
                )
            if validation.allow_blank:
                parts.append("allow_blank=true")
            output.write("- " + "; ".join(parts) + "\n")

        if len(validations) > self.MAX_DATA_VALIDATIONS:
            output.write(
                f"[Data validations truncated: "
                f"{len(validations) - self.MAX_DATA_VALIDATIONS} omitted]\n"
            )

    def _write_conditional_formats(self, output: TextIO, worksheet: Any) -> None:
        """Handle write conditional formats."""
        conditional = worksheet.conditional_formatting
        items = list(conditional)
        if not items:
            return

        output.write("Conditional formatting:\n")
        for item in items[: self.MAX_CONDITIONAL_FORMATS]:
            ranges = getattr(item, "sqref", item)
            try:
                rules = conditional[item]
            except Exception:
                rules = ()
            rule_types = [str(getattr(rule, "type", "unknown")) for rule in rules]
            suffix = "; rules=" + ",".join(rule_types) if rule_types else ""
            output.write(f"- range={ranges}{suffix}\n")

        if len(items) > self.MAX_CONDITIONAL_FORMATS:
            output.write(
                f"[Conditional formatting truncated: "
                f"{len(items) - self.MAX_CONDITIONAL_FORMATS} omitted]\n"
            )

    def _material_cells(self, worksheet: Any) -> list[Any]:
        """Handle material cells."""
        sparse_store = getattr(worksheet, "_cells", None)
        if isinstance(sparse_store, dict):
            cells = [
                cell
                for cell in sparse_store.values()
                if not isinstance(cell, MergedCell) and self._is_material_cell(cell)
            ]
            cells.sort(key=lambda cell: (cell.row, cell.column))
            return cells

        # Public fallback if openpyxl changes its normal-mode sparse store.
        # This fallback is intentionally bounded rather than trusting a possibly
        # inflated worksheet dimension.
        max_row = min(worksheet.max_row, self.MAX_CELLS_PER_SHEET)
        max_col = min(worksheet.max_column, 256)
        cells: list[Any] = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
        ):
            for cell in row:
                if self._is_material_cell(cell):
                    cells.append(cell)
                    if len(cells) >= self.MAX_CELLS_PER_SHEET:
                        return cells
        return cells

    @staticmethod
    def _is_material_cell(cell: Any) -> bool:
        """Handle is material cell."""
        return (
            cell.value is not None
            or cell.comment is not None
            or cell.hyperlink is not None
        )

    def _render_cell(self, cell: Any, cached_cell: Any | None) -> str:
        """Handle render cell."""
        value = cell.value

        if isinstance(value, ArrayFormula):
            rendered = self._bounded_inline(
                value.text or "[array formula text unavailable]",
                self.MAX_CELL_CHARS,
            ) + f" [array ref={value.ref}]"
            rendered = self._append_cached(rendered, cached_cell)
        elif isinstance(value, DataTableFormula):
            attrs = ", ".join(f"{key}={val}" for key, val in dict(value).items())
            rendered = f"[data-table formula: {attrs}]"
            rendered = self._append_cached(rendered, cached_cell)
        elif cell.data_type == "f":
            rendered = self._bounded_inline(value, self.MAX_CELL_CHARS)
            rendered = self._append_cached(rendered, cached_cell)
        else:
            rendered = self._render_scalar(value, max_chars=self.MAX_CELL_CHARS)

        annotations: list[str] = []
        if cell.number_format and cell.number_format != "General":
            annotations.append(f"format={self._quote(cell.number_format)}")
        if cell.style and cell.style != "Normal":
            annotations.append(f"style={self._quote(cell.style)}")

        if cell.hyperlink is not None:
            hyperlink = cell.hyperlink
            target = (
                getattr(hyperlink, "target", None)
                or getattr(hyperlink, "location", None)
                or str(hyperlink)
            )
            annotations.append(
                "hyperlink="
                + self._quote(
                    self._bounded_inline(target, self.MAX_ANNOTATION_CHARS)
                )
            )

        if cell.comment is not None:
            author = getattr(cell.comment, "author", None)
            text = self._bounded_inline(
                getattr(cell.comment, "text", ""), self.MAX_ANNOTATION_CHARS
            )
            if author:
                text = f"{author}: {text}"
            annotations.append("comment=" + self._quote(text))

        if annotations:
            rendered += " {" + "; ".join(annotations) + "}"
        return rendered

    def _append_cached(self, rendered: str, cached_cell: Any | None) -> str:
        """Handle append cached."""
        if cached_cell is None:
            return rendered + " [cached value unavailable]"
        if cached_cell.value is None:
            return rendered + " [cached value unavailable-or-blank]"
        return (
            rendered
            + " [cached="
            + self._render_scalar(cached_cell.value, max_chars=self.MAX_CELL_CHARS)
            + "]"
        )

    def _render_scalar(self, value: Any, *, max_chars: int | None = None) -> str:
        """Handle render scalar."""
        if value is None:
            text = "null"
        elif isinstance(value, bool):
            text = "TRUE" if value else "FALSE"
        elif isinstance(value, datetime):
            text = value.isoformat(sep="T")
        elif isinstance(value, (date, time)):
            text = value.isoformat()
        elif isinstance(value, float):
            if math.isnan(value):
                text = "NaN"
            elif math.isinf(value):
                text = "Infinity" if value > 0 else "-Infinity"
            else:
                text = repr(value)
        elif isinstance(value, str):
            text = json.dumps(value, ensure_ascii=False)
        else:
            # CellRichText and normal numeric scalar types stringify well.
            text = str(value)

        return self._bounded_inline(text, max_chars) if max_chars else text

    @staticmethod
    def _quote(value: Any) -> str:
        """Handle quote."""
        return json.dumps(str(value), ensure_ascii=False)

    @staticmethod
    def _bounded_inline(value: Any, limit: int) -> str:
        """Handle bounded inline."""
        text = (
            str(value)
            .replace("\r\n", "\n")
            .replace("\r", "\n")
            .replace("\t", "\\t")
            .replace("\n", "\\n")
        )
        if len(text) <= limit:
            return text
        omitted = len(text) - limit
        return text[:limit] + f"…[truncated {omitted} chars]"

    @staticmethod
    def _write_bounded_list(
        output: TextIO,
        label: str,
        values: list[str],
        limit: int,
    ) -> None:
        """Handle write bounded list."""
        if not values:
            return
        visible = values[:limit]
        output.write(f"{label}: " + ", ".join(visible) + "\n")
        if len(values) > len(visible):
            output.write(
                f"[{label} truncated: {len(values) - len(visible)} omitted]\n"
            )
